import argparse
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from dcf_engine import DEFAULT_PROJECTION_SPECS


DEFAULT_WORKBOOK_CANDIDATES = (
    "Stock Pitch - mine Excel Model-1.xlsx",
    "Stock Pitch - mine Excel Model.xlsx",
)


def _set_percent_range(ws, cells: list[str]) -> None:
    for ref in cells:
        ws[ref].number_format = "0.0%"


def _set_currency_range(ws, cells: list[str]) -> None:
    for ref in cells:
        ws[ref].number_format = "$#,##0.00"


def _resolve_workbook_path(workbook: str | None) -> Path:
    if workbook:
        path = Path(workbook)
        if not path.exists():
            raise FileNotFoundError(path)
        return path

    for candidate in DEFAULT_WORKBOOK_CANDIDATES:
        path = Path(candidate)
        if path.exists():
            return path

    raise FileNotFoundError(
        "No workbook found. Pass --workbook or place a stock-pitch workbook in this directory."
    )


def main(workbook: str | None = None) -> None:
    workbook_path = _resolve_workbook_path(workbook)
    wb = openpyxl.load_workbook(workbook_path)

    op = wb["OP"]
    wacc = wb["WACC"]
    dcf = wb["DCF Model"]
    if "MC Assumptions" in wb.sheetnames:
        del wb["MC Assumptions"]
    mc = wb.create_sheet("MC Assumptions")

    # Replace broken external references with internal links.
    wacc["G6"] = "=OP!U11"
    wacc["G11"] = "=OP!U17"

    # Wire the operating page to the DCF output.
    op["U8"] = "='DCF Model'!D83"
    op["U9"] = '=IFERROR(U8/U6-1,"")'

    # Preserve user-entered WACC on the DCF sheet; only keep the live share-price link.
    dcf["D85"] = "=OP!U6"
    dcf["D86"] = '=IFERROR(D83/D85-1,"")'

    projected_cols = range(10, 15)  # J:N
    for col in projected_cols:
        letter = get_column_letter(col)
        dcf[f"{letter}19"] = f"={letter}10-{letter}13-{letter}16"
        dcf[f"{letter}20"] = f'=IFERROR({letter}19/{letter}7,"")'
        dcf[f"{letter}50"] = f"={letter}44+{letter}46+{letter}47-{letter}48"
        dcf[f"{letter}53"] = f'=IFERROR({letter}50/(1+$D$57)^{letter}52,"")'

    dcf["D61"] = '=IFERROR($N$50*(1+$D$58)/($D$57-$D$58),"")'
    dcf["D62"] = '=IFERROR(($N$19+$N$24)*$D$59,"")'
    dcf["D63"] = '=IF($D$64=1,$D$61,$D$62)'
    dcf["D66"] = '=IFERROR($D$63/(1+$D$57)^$N$52,"")'
    dcf["D70"] = '=IFERROR(SUM($J$53:$N$53),"")'
    dcf["D71"] = "=$D$66"
    dcf["D73"] = "=$D$70+$D$71"
    dcf["D80"] = "=$D$73+$D$75-$D$76-$D$77-$D$78"
    dcf["D83"] = '=IFERROR($D$80/$D$82,"")'

    # WACC vs terminal growth sensitivity.
    dcf["D91"] = "=$D$58-1.5%"
    for col in range(5, 11):
        prev = get_column_letter(col - 1)
        curr = get_column_letter(col)
        dcf[f"{curr}91"] = f"={prev}91+0.5%"

    dcf["C92"] = "=$D$57-1.5%"
    for row in range(93, 99):
        dcf[f"C{row}"] = f"=C{row-1}+0.5%"

    for row in range(92, 99):
        for col in range(4, 11):
            letter = get_column_letter(col)
            dcf[f"{letter}{row}"] = (
                f'=IFERROR(('
                f'SUMPRODUCT($J$50:$N$50/((1+$C{row})^$J$52:$N$52))'
                f'+(($N$50*(1+{letter}$91))/($C{row}-{letter}$91))/((1+$C{row})^$N$52)'
                f'+$D$75-$D$76-$D$77-$D$78'
                f')/$D$82,"")'
            )

    # WACC vs exit multiple sensitivity.
    dcf["D102"] = "=$D$59-3"
    for col in range(5, 11):
        prev = get_column_letter(col - 1)
        curr = get_column_letter(col)
        dcf[f"{curr}102"] = f"={prev}102+1"

    dcf["C103"] = "=$D$57-1.5%"
    for row in range(104, 110):
        dcf[f"C{row}"] = f"=C{row-1}+0.5%"

    for row in range(103, 110):
        for col in range(4, 11):
            letter = get_column_letter(col)
            dcf[f"{letter}{row}"] = (
                f'=IFERROR(('
                f'SUMPRODUCT($J$50:$N$50/((1+$C{row})^$J$52:$N$52))'
                f'+((($N$19+$N$24)*{letter}$102)/((1+$C{row})^$N$52))'
                f'+$D$75-$D$76-$D$77-$D$78'
                f')/$D$82,"")'
            )

    _set_percent_range(
        dcf,
        ["D57", "D58", "D86"]
        + [f"{get_column_letter(col)}91" for col in range(4, 11)]
        + [f"C{row}" for row in range(92, 99)]
        + [f"C{row}" for row in range(103, 110)],
    )
    _set_currency_range(
        dcf,
        ["D83", "D85"]
        + [f"{get_column_letter(col)}{row}" for row in range(92, 99) for col in range(4, 11)]
        + [f"{get_column_letter(col)}{row}" for row in range(103, 110) for col in range(4, 11)],
    )
    for col in range(4, 11):
        dcf[f"{get_column_letter(col)}102"].number_format = "0.0x"

    op["U8"].number_format = "$#,##0.00"
    op["U9"].number_format = "0.0%"
    dcf["D59"].number_format = "0.0x"

    headers = ["Driver", "Label", "Shock Kind", "Spread", "Low", "High", "Per Year?", "Description"]
    descriptions = {
        "revenue": "Relative shock applied to projected revenue by year.",
        "cogs_pct": "Additive shock to COGS as a percent of revenue.",
        "rd_pct": "Additive shock to R&D as a percent of revenue.",
        "sga_pct": "Additive shock to SG&A as a percent of revenue.",
        "da_pct": "Additive shock to D&A as a percent of revenue.",
        "tax_rate": "Additive shock to the effective tax rate.",
        "capex": "Relative shock applied to capital expenditures.",
        "nwc": "Relative shock applied to change in working capital.",
        "wacc": "Scalar shock to the discount rate.",
        "exit_multiple": "Scalar shock to the exit EV/EBITDA multiple.",
        "terminal_g": "Scalar shock to the terminal growth rate.",
    }
    for col, header in enumerate(headers, start=1):
        mc.cell(row=1, column=col, value=header)
    for row, spec in enumerate(DEFAULT_PROJECTION_SPECS, start=2):
        mc.cell(row=row, column=1, value=spec.name)
        mc.cell(row=row, column=2, value=spec.label)
        mc.cell(row=row, column=3, value=spec.kind)
        mc.cell(row=row, column=4, value=spec.spread)
        mc.cell(row=row, column=5, value=spec.low)
        mc.cell(row=row, column=6, value=spec.high)
        mc.cell(row=row, column=7, value="Y" if spec.applies_per_year else "N")
        mc.cell(row=row, column=8, value=descriptions[spec.name])
    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        mc.column_dimensions[col].width = 18 if col != "H" else 42

    # Ask spreadsheet apps to recalc on open.
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.calcMode = "auto"
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True

    wb.save(workbook_path)
    print(f"Updated workbook: {workbook_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Patch the stock-pitch workbook for Monte Carlo integration.")
    parser.add_argument(
        "--workbook",
        default=None,
        help="Workbook path. If omitted, the script auto-detects the stock-pitch workbook in this directory.",
    )
    args = parser.parse_args()
    main(args.workbook)

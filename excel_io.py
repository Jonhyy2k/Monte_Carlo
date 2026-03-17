"""
Phase 6 — Excel I/O Stub

Creates a dummy Excel workbook with an Inputs sheet.
Provides read_inputs_from_excel() and write_results_to_excel().
"""

from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage

from dcf_engine import (
    DEFAULT_PROJECTION_SPECS,
    ProjectionShockSpec,
    run_base_dcf_from_imported_data,
)


# ---------------------------------------------------------------------------
# Cell map: parameter name → (sheet, cell)
# ---------------------------------------------------------------------------
INPUT_CELL_MAP = {
    "revenue_base":         ("Inputs", "B2"),
    "revenue_growth":       ("Inputs", "B3"),
    "projection_years":     ("Inputs", "B4"),
    "cogs_pct":             ("Inputs", "B5"),
    "sga_pct":              ("Inputs", "B6"),
    "capex_intensity":      ("Inputs", "B7"),
    "tax_rate":             ("Inputs", "B8"),
    "exit_multiple":        ("Inputs", "B9"),
    "terminal_g":           ("Inputs", "B10"),
    "wacc":                 ("Inputs", "B11"),
    "net_debt":             ("Inputs", "B12"),
    "shares_outstanding":   ("Inputs", "B13"),
    "exit_multiple_weight": ("Inputs", "B14"),
    "current_price":        ("Inputs", "B15"),
    "fundamental_target":   ("Inputs", "B16"),
}

# Default placeholder values (generic company)
DEFAULTS = {
    "revenue_base":         10_000.0,
    "revenue_growth":       0.08,
    "projection_years":     5,
    "cogs_pct":             0.55,
    "sga_pct":              0.15,
    "capex_intensity":      0.06,
    "tax_rate":             0.25,
    "exit_multiple":        12.0,
    "terminal_g":           0.025,
    "wacc":                 0.10,
    "net_debt":             2_000.0,
    "shares_outstanding":   500.0,
    "exit_multiple_weight": 0.50,
    "current_price":        55.0,
    "fundamental_target":   68.0,
}


DCF_LINE_ITEMS = [
    "Revenue", "COGS", "Gross Profit", "R&D", "SG&A", "EBITDA",
    "D&A", "EBIT", "Taxes", "NOPAT", "CapEx", "Chg in NWC", "FCFF",
]

IMPORT_SHEET_CANDIDATES = ("ImportedDCF", "DCF", "DCF Model")
MC_ASSUMPTIONS_SHEET = "MC Assumptions"

SERIES_ALIASES = {
    "Revenue": ("revenue", "sales"),
    "COGS": ("cogs", "cost of goods sold"),
    "COGS %": ("cogs %", "cogs pct", "cogs percentage", "cost of goods sold %"),
    "R&D": ("r&d", "research & development", "research and development"),
    "R&D %": ("r&d %", "r&d pct", "research & development %", "research and development %"),
    "SG&A": ("sg&a", "sga", "selling general & administrative"),
    "SG&A %": ("sg&a %", "sga %", "sg&a pct", "sga pct"),
    "D&A": ("d&a", "da", "depreciation & amortization", "depreciation and amortization"),
    "D&A %": ("d&a %", "da %", "d&a pct", "da pct"),
    "Taxes": ("taxes", "cash taxes"),
    "Tax Rate": ("tax rate", "effective tax rate"),
    "EBIT": ("ebit",),
    "CapEx": ("capex", "capital expenditures", "capital expenditure"),
    "CapEx %": ("capex %", "capex pct", "capital expenditures %"),
    "Chg in NWC": ("chg in nwc", "change in nwc", "delta nwc", "change in working capital"),
    "FCFF": ("fcff", "free cash flow to firm"),
}

DCF_SCALAR_ASSUMPTIONS = {
    "WACC":                 0.10,
    "Exit Multiple":        12.0,
    "Terminal g":           0.025,
    "Net Debt":             2_000.0,
    "Shares Outstanding":   500.0,
    "Current Price":        55.0,
    "Target Price":         68.0,
    "Exit Multiple Weight": 0.50,
}


def _build_dcf_dummy_data(n_years: int = 5) -> dict[str, list[float]]:
    """Build a small importable DCF export for the Monte Carlo layer."""
    rev_base = DEFAULTS["revenue_base"]
    g = DEFAULTS["revenue_growth"]
    cogs_pct = DEFAULTS["cogs_pct"]
    rd_pct = 0.07
    sga_pct = max(DEFAULTS["sga_pct"] - rd_pct, 0.03)
    capex_pct = DEFAULTS["capex_intensity"]
    tax_rate = DEFAULTS["tax_rate"]

    data = {item: [] for item in DCF_LINE_ITEMS}
    for t in range(n_years):
        rev = rev_base * (1 + g) ** (t + 1)
        cogs = -rev * cogs_pct
        gross = rev + cogs
        rd = -rev * rd_pct
        sga = -rev * sga_pct
        ebitda = gross + rd + sga
        da = -rev * capex_pct
        ebit = ebitda + da
        taxes = -(abs(ebit) * tax_rate)
        nopat = ebit + taxes
        capex = -rev * capex_pct
        nwc = -rev * 0.01  # small NWC change
        fcff = nopat - da + capex + nwc  # NOPAT + D&A - CapEx - Chg NWC

        data["Revenue"].append(round(rev, 1))
        data["COGS"].append(round(cogs, 1))
        data["Gross Profit"].append(round(gross, 1))
        data["R&D"].append(round(rd, 1))
        data["SG&A"].append(round(sga, 1))
        data["EBITDA"].append(round(ebitda, 1))
        data["D&A"].append(round(da, 1))
        data["EBIT"].append(round(ebit, 1))
        data["Taxes"].append(round(taxes, 1))
        data["NOPAT"].append(round(nopat, 1))
        data["CapEx"].append(round(capex, 1))
        data["Chg in NWC"].append(round(nwc, 1))
        data["FCFF"].append(round(fcff, 1))
    return data


def create_stub_excel(filepath: str = "model_inputs.xlsx") -> None:
    """Create the legacy stub workbook used only when no real DCF workbook is present."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inputs"

    # Header
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws["B1"].font = openpyxl.styles.Font(bold=True)

    for param, (sheet, cell) in INPUT_CELL_MAP.items():
        row = int(cell[1:])
        ws[f"A{row}"] = param
        ws[cell] = DEFAULTS[param]

    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    # ---- Imported DCF export sheet ----
    n_years = int(DEFAULTS["projection_years"])
    dcf_data = _build_dcf_dummy_data(n_years)
    bold = openpyxl.styles.Font(bold=True)

    dcf = wb.create_sheet("ImportedDCF")
    dcf["A1"] = "Line Item"
    dcf["A1"].font = bold
    for c in range(n_years):
        cell = dcf.cell(row=1, column=c + 2, value=f"Year {c + 1}")
        cell.font = bold

    for r, item in enumerate(DCF_LINE_ITEMS, start=2):
        dcf.cell(row=r, column=1, value=item)
        for c, val in enumerate(dcf_data[item]):
            dcf.cell(row=r, column=c + 2, value=val)

    scalar_start = len(DCF_LINE_ITEMS) + 3
    dcf.cell(row=scalar_start, column=1, value="Assumptions").font = bold
    for i, (key, val) in enumerate(DCF_SCALAR_ASSUMPTIONS.items()):
        r = scalar_start + 1 + i
        dcf.cell(row=r, column=1, value=key)
        dcf.cell(row=r, column=2, value=val)

    dcf.column_dimensions["A"].width = 22
    for c in range(n_years):
        dcf.column_dimensions[openpyxl.utils.get_column_letter(c + 2)].width = 14

    # ---- Monte Carlo shock assumptions ----
    mc = wb.create_sheet(MC_ASSUMPTIONS_SHEET)
    headers = ["Driver", "Label", "Shock Kind", "Spread", "Low", "High", "Per Year?", "Description"]
    for col, header in enumerate(headers, start=1):
        mc.cell(row=1, column=col, value=header).font = bold

    descriptions = {
        "revenue": "Relative shock applied to imported revenue by year.",
        "cogs_pct": "Additive margin shock to imported COGS / revenue.",
        "rd_pct": "Additive margin shock to imported R&D / revenue.",
        "sga_pct": "Additive margin shock to imported SG&A / revenue.",
        "da_pct": "Additive margin shock to imported D&A / revenue.",
        "tax_rate": "Additive shock to imported operating tax rate.",
        "capex": "Relative shock applied to imported CapEx.",
        "nwc": "Relative shock applied to imported change in NWC.",
        "wacc": "Scalar shock to discount rate.",
        "exit_multiple": "Scalar shock to exit multiple.",
        "terminal_g": "Scalar shock to terminal growth.",
    }
    for row, spec in enumerate(DEFAULT_PROJECTION_SPECS, start=2):
        mc.cell(row=row, column=1, value=spec.name)
        mc.cell(row=row, column=2, value=spec.label)
        mc.cell(row=row, column=3, value=spec.kind)
        mc.cell(row=row, column=4, value=spec.spread)
        mc.cell(row=row, column=5, value=spec.low)
        mc.cell(row=row, column=6, value=spec.high)
        mc.cell(row=row, column=7, value="Y" if spec.applies_per_year else "N")
        mc.cell(row=row, column=8, value=descriptions.get(spec.name, ""))

    for col in ("A", "B", "C", "D", "E", "F", "G", "H"):
        mc.column_dimensions[col].width = 18 if col != "H" else 42

    wb.save(filepath)
    print(f"  Created stub: {filepath}")


def _normalize_label(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def _find_import_sheet_name(wb: openpyxl.Workbook) -> str | None:
    for sheet_name in IMPORT_SHEET_CANDIDATES:
        if sheet_name in wb.sheetnames:
            return sheet_name
    return None


def _row_to_array(ws, row: int, n_years: int) -> np.ndarray:
    values = []
    for col in range(2, 2 + n_years):
        cell_value = ws.cell(row=row, column=col).value
        values.append(float(cell_value) if cell_value is not None else 0.0)
    return np.array(values, dtype=float)


def _lookup_series(row_map: dict[str, np.ndarray], canonical_name: str) -> np.ndarray | None:
    for alias in SERIES_ALIASES.get(canonical_name, ()):
        if alias in row_map:
            return row_map[alias]
    return None


def _standardize_nwc(raw: np.ndarray) -> np.ndarray:
    negatives = int(np.count_nonzero(raw < 0))
    positives = int(np.count_nonzero(raw > 0))
    if negatives >= positives:
        return -raw
    return raw


def read_projection_specs_from_excel(filepath: str = "model_inputs.xlsx") -> list[ProjectionShockSpec]:
    """Read stochastic shock specs from the workbook, falling back to defaults."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if MC_ASSUMPTIONS_SHEET not in wb.sheetnames:
        wb.close()
        return list(DEFAULT_PROJECTION_SPECS)

    ws = wb[MC_ASSUMPTIONS_SHEET]
    specs: list[ProjectionShockSpec] = []
    defaults = {spec.name: spec for spec in DEFAULT_PROJECTION_SPECS}

    for row in range(2, ws.max_row + 1):
        name = _normalize_label(ws.cell(row=row, column=1).value)
        if not name:
            continue
        default = defaults.get(name)
        if default is None:
            continue

        kind = str(ws.cell(row=row, column=3).value or default.kind).strip().lower()
        spread = ws.cell(row=row, column=4).value
        low = ws.cell(row=row, column=5).value
        high = ws.cell(row=row, column=6).value
        per_year_value = _normalize_label(ws.cell(row=row, column=7).value)
        applies_per_year = per_year_value in {"y", "yes", "true", "1"}

        specs.append(
            ProjectionShockSpec(
                name=default.name,
                label=str(ws.cell(row=row, column=2).value or default.label),
                kind=kind,
                spread=float(spread) if spread is not None else default.spread,
                low=float(low) if low is not None else default.low,
                high=float(high) if high is not None else default.high,
                applies_per_year=applies_per_year,
            )
        )

    wb.close()

    if not specs:
        return list(DEFAULT_PROJECTION_SPECS)

    spec_map = {spec.name: spec for spec in specs}
    ordered = []
    for default in DEFAULT_PROJECTION_SPECS:
        ordered.append(spec_map.get(default.name, default))
    return ordered


def read_dcf_from_excel(filepath: str = "model_inputs.xlsx") -> dict:
    """
    Read the imported DCF export sheet → standardized yearly arrays + scalars.

    Returns
    -------
    dict with keys:
        "arrays"  → {line_item_name: np.array of yearly values}
        "n_years" → int
        "wacc", "exit_multiple", "terminal_g", "net_debt",
        "shares_outstanding", "current_price", "target_price",
        "exit_multiple_weight" → floats
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = _find_import_sheet_name(wb)
    if sheet_name is None:
        wb.close()
        raise KeyError("Workbook does not contain an ImportedDCF or DCF sheet")
    if sheet_name == "DCF Model":
        ws_data = wb["DCF Model"]
        op_data = wb["OP"]
        wacc_data = wb["WACC"]
        wb_formula = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb_formula["DCF Model"]
        op = wb_formula["OP"]
        wacc_ws = wb_formula["WACC"]
        projected_cols = range(10, 15)  # J:N

        def _require_float(value, label: str) -> float:
            if isinstance(value, (int, float)):
                return float(value)
            raise ValueError(f"{label} must be a numeric value in the workbook")

        def _maybe_float(value) -> float | None:
            if isinstance(value, (int, float)):
                return float(value)
            return None

        def _value(data_sheet, formula_sheet, ref: str, label: str) -> float:
            data_value = data_sheet[ref].value
            if isinstance(data_value, (int, float)):
                return float(data_value)
            return _require_float(formula_sheet[ref].value, label)

        def _literal_series(row: int) -> list[float | None]:
            values = []
            for col in projected_cols:
                data_value = ws_data.cell(row=row, column=col).value
                formula_value = ws.cell(row=row, column=col).value
                if isinstance(data_value, (int, float)):
                    values.append(float(data_value))
                elif isinstance(formula_value, (int, float)):
                    values.append(float(formula_value))
                else:
                    values.append(None)
            return values

        def _filled_series(row: int, label: str) -> np.ndarray:
            raw = _literal_series(row)
            if any(value is None for value in raw):
                raise ValueError(f"{label} projected assumptions must be numeric in DCF Model")
            return np.array(raw, dtype=float)

        hist_revenue = _value(ws_data, ws, "H7", "Historical revenue")
        hist_growth_g = _value(ws_data, ws, "G8", "Historical growth (2024)")
        hist_growth_h = _value(ws_data, ws, "H8", "Historical growth (2025)")

        growth_inputs = _literal_series(8)
        growth = np.empty(len(projected_cols), dtype=float)
        growth[0] = growth_inputs[0] if growth_inputs[0] is not None else (hist_growth_g + hist_growth_h) / 2
        for i in range(1, len(projected_cols)):
            growth[i] = growth_inputs[i] if growth_inputs[i] is not None else growth[i - 1] + 0.005

        revenue_inputs = _literal_series(7)
        revenue = np.empty(len(projected_cols), dtype=float)
        revenue[0] = revenue_inputs[0] if revenue_inputs[0] is not None else hist_revenue * (1 + growth[0])
        for i in range(1, len(projected_cols)):
            revenue[i] = revenue_inputs[i] if revenue_inputs[i] is not None else revenue[i - 1] * (1 + growth[i])

        gross_margin = _filled_series(11, "Gross margin")
        rd_pct = _filled_series(14, "R&D %")
        sga_other_pct = _filled_series(17, "SG&A %")
        da_pct = _filled_series(25, "D&A %")
        capex_pct = _filled_series(31, "CapEx %")
        nwc_pct = _filled_series(34, "Change in NWC % of delta revenue")
        tax_rate = np.clip(_filled_series(42, "Tax rate"), 0.0, 0.45)

        cogs = revenue * (1 - gross_margin)
        rd = revenue * rd_pct
        sga = revenue * sga_other_pct
        da = revenue * da_pct
        capex = np.abs(revenue * capex_pct)
        delta_revenue = np.empty(len(projected_cols), dtype=float)
        delta_revenue[0] = revenue[0] - hist_revenue
        delta_revenue[1:] = revenue[1:] - revenue[:-1]
        nwc = nwc_pct * delta_revenue

        cash = _value(ws_data, ws, "D75", "Cash")
        total_debt = _value(ws_data, ws, "D76", "Total debt")
        preferred = _value(ws_data, ws, "D77", "Preferred stock")
        minority = _value(ws_data, ws, "D78", "Minority interest")
        shares = _value(ws_data, ws, "D82", "Shares outstanding")
        current_price = _maybe_float(ws_data["D85"].value)
        if current_price is None:
            current_price = _maybe_float(op_data["U6"].value)
        if current_price is None:
            current_price = _require_float(op["U6"].value, "Current share price")
        terminal_g = _value(ws_data, ws, "D58", "Terminal growth rate")
        exit_multiple = _value(ws_data, ws, "D59", "Exit multiple")
        tv_method = int(_value(ws_data, ws, "D64", "Terminal value method"))
        discount_periods = _filled_series(52, "Discount periods")

        wacc = _maybe_float(ws_data["D57"].value)
        if wacc is None:
            wacc = _maybe_float(ws["D57"].value)
        if wacc is None:
            market_cap = _value(op_data, op, "U11", "Market cap")
            beta = _value(op_data, op, "U17", "Beta")
            debt = _value(wacc_data, wacc_ws, "G15", "Debt")
            risk_free = _value(wacc_data, wacc_ws, "G9", "Risk-free rate")
            equity_risk_premium = _value(wacc_data, wacc_ws, "G12", "Equity risk premium")
            tax_shield = _value(wacc_data, wacc_ws, "G18", "Tax rate for WACC")

            cost_of_debt_value = wacc_data["G17"].value
            if isinstance(cost_of_debt_value, (int, float)):
                cost_of_debt = float(cost_of_debt_value)
            else:
                formula = str(wacc_ws["G17"].value or "")
                if formula.startswith("=") and "/G15" in formula:
                    interest_expense = float(formula[1:].split("/")[0])
                    cost_of_debt = interest_expense / debt
                else:
                    raise ValueError("Cost of debt must be numeric or a simple interest-expense / debt formula")

            total_capital = market_cap + debt
            cost_of_equity = risk_free + beta * equity_risk_premium
            wacc = (
                market_cap / total_capital * cost_of_equity
                + debt / total_capital * cost_of_debt * (1 - tax_shield)
            )

        data = {
            "sheet_name": sheet_name,
            "arrays": {
                "Revenue": revenue,
                "COGS": cogs,
                "R&D": rd,
                "SG&A": sga,
                "D&A": da,
                "CapEx": capex,
                "Chg in NWC": nwc,
                "Tax Rate": tax_rate,
            },
            "n_years": len(projected_cols),
            "wacc": wacc,
            "exit_multiple": exit_multiple,
            "terminal_g": terminal_g,
            "net_debt": total_debt + preferred + minority - cash,
            "shares_outstanding": shares,
            "current_price": current_price,
            "exit_multiple_weight": 0.0 if tv_method == 1 else 1.0,
            "discount_periods": discount_periods,
            "shock_specs": read_projection_specs_from_excel(filepath),
        }
        data["target_price"] = run_base_dcf_from_imported_data(data)
        wb_formula.close()
        wb.close()
        return data
    dcf = wb[sheet_name]

    n_years = 0
    for c in range(2, 50):
        if dcf.cell(row=1, column=c).value is None:
            break
        n_years += 1

    row_map = {}
    row = 2
    while True:
        label = dcf.cell(row=row, column=1).value
        normalized = _normalize_label(label)
        if label is None or normalized == "assumptions":
            break
        if normalized:
            row_map[normalized] = _row_to_array(dcf, row, n_years)
        row += 1

    scalar_map = {
        "WACC": "wacc",
        "Exit Multiple": "exit_multiple",
        "Terminal g": "terminal_g",
        "Net Debt": "net_debt",
        "Shares Outstanding": "shares_outstanding",
        "Current Price": "current_price",
        "Target Price": "target_price",
        "Exit Multiple Weight": "exit_multiple_weight",
    }
    scalars = {}
    for r in range(row, dcf.max_row + 1):
        label = dcf.cell(row=r, column=1).value
        if label in scalar_map:
            val = dcf.cell(row=r, column=2).value
            scalars[scalar_map[label]] = float(val) if val is not None else DCF_SCALAR_ASSUMPTIONS[label]

    for excel_name, key in scalar_map.items():
        if key not in scalars:
            scalars[key] = DCF_SCALAR_ASSUMPTIONS[excel_name]

    revenue = _lookup_series(row_map, "Revenue")
    if revenue is None:
        wb.close()
        raise KeyError("Imported DCF sheet must include a Revenue row")
    revenue = np.asarray(revenue, dtype=float)

    def _expense_or_ratio(
        expense_name: str,
        ratio_name: str,
        allow_missing: bool = False,
    ) -> np.ndarray:
        absolute = _lookup_series(row_map, expense_name)
        ratio = _lookup_series(row_map, ratio_name)
        if absolute is not None:
            return np.abs(absolute)
        if ratio is not None:
            return revenue * np.clip(np.asarray(ratio, dtype=float), 0.0, 2.0)
        if allow_missing:
            return np.zeros(n_years, dtype=float)
        raise KeyError(f"Imported DCF sheet is missing {expense_name} or {ratio_name}")

    cogs = _expense_or_ratio("COGS", "COGS %")
    rd = _expense_or_ratio("R&D", "R&D %", allow_missing=True)
    sga = _expense_or_ratio("SG&A", "SG&A %")
    da = _expense_or_ratio("D&A", "D&A %")
    capex = _expense_or_ratio("CapEx", "CapEx %")

    tax_rate = _lookup_series(row_map, "Tax Rate")
    if tax_rate is None:
        taxes = _lookup_series(row_map, "Taxes")
        ebit = _lookup_series(row_map, "EBIT")
        if taxes is None or ebit is None:
            tax_rate = np.full(n_years, DEFAULTS["tax_rate"], dtype=float)
        else:
            tax_rate = np.clip(np.abs(taxes) / np.maximum(np.abs(ebit), 1e-9), 0.0, 0.45)
    else:
        tax_rate = np.clip(np.asarray(tax_rate, dtype=float), 0.0, 0.45)

    nwc_raw = _lookup_series(row_map, "Chg in NWC")
    if nwc_raw is None:
        nwc = np.zeros(n_years, dtype=float)
    else:
        nwc = np.asarray(_standardize_nwc(np.asarray(nwc_raw, dtype=float)), dtype=float)

    wb.close()
    return {
        "sheet_name": sheet_name,
        "arrays": {
            "Revenue": revenue,
            "COGS": cogs,
            "R&D": rd,
            "SG&A": sga,
            "D&A": da,
            "CapEx": capex,
            "Chg in NWC": nwc,
            "Tax Rate": tax_rate,
        },
        "n_years": n_years,
        "shock_specs": read_projection_specs_from_excel(filepath),
        **scalars,
    }


def has_dcf_sheet(filepath: str = "model_inputs.xlsx") -> bool:
    """Check if the workbook contains an importable DCF sheet."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    result = _find_import_sheet_name(wb) is not None
    wb.close()
    return result


def read_inputs_from_excel(filepath: str = "model_inputs.xlsx") -> dict:
    """
    Read model inputs from Excel and return a dictionary.

    When the real model arrives, just update INPUT_CELL_MAP references.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    if "Inputs" not in wb.sheetnames:
        has_import_sheet = _find_import_sheet_name(wb) is not None
        wb.close()
        if not has_import_sheet:
            raise KeyError("Workbook does not contain an Inputs sheet or an importable DCF sheet")

        dcf_data = read_dcf_from_excel(filepath)
        revenue = np.asarray(dcf_data["arrays"]["Revenue"], dtype=float)
        return {
            **DEFAULTS,
            "revenue_base": float(revenue[0]) if revenue.size else DEFAULTS["revenue_base"],
            "projection_years": int(dcf_data["n_years"]),
            "exit_multiple": float(dcf_data["exit_multiple"]),
            "terminal_g": float(dcf_data["terminal_g"]),
            "wacc": float(dcf_data["wacc"]),
            "net_debt": float(dcf_data["net_debt"]),
            "shares_outstanding": float(dcf_data["shares_outstanding"]),
            "exit_multiple_weight": float(dcf_data["exit_multiple_weight"]),
            "current_price": float(dcf_data.get("current_price", DEFAULTS["current_price"])),
            "fundamental_target": float(dcf_data.get("target_price", DEFAULTS["fundamental_target"])),
        }

    inputs = {}
    for param, (sheet, cell) in INPUT_CELL_MAP.items():
        ws = wb[sheet]
        val = ws[cell].value
        if val is None:
            val = DEFAULTS.get(param)
        inputs[param] = val
    wb.close()
    return inputs


def write_results_to_excel(
    filepath: str = "model_inputs.xlsx",
    results: dict | None = None,
    chart_paths: dict[str, Path | str] | None = None,
) -> None:
    """
    Write Monte Carlo output stats into a Results sheet.

    Expected keys in results dict:
        mean, median, std, p5, p25, p50, p75, p95,
        n_iterations, n_valid, elapsed_seconds,
        sobol_df (optional pandas DataFrame)
    """
    results = results or {}
    path = Path(filepath)
    if path.exists():
        wb = openpyxl.load_workbook(filepath)
    else:
        wb = openpyxl.Workbook()

    # Remove old Results sheet if present
    if "Results" in wb.sheetnames:
        del wb["Results"]
    ws = wb.create_sheet("Results")

    bold = openpyxl.styles.Font(bold=True)
    title_font = openpyxl.styles.Font(bold=True, size=16)
    section_font = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
    money_fmt = "$#,##0.00"
    pct_fmt = "0.0%"
    int_fmt = "#,##0"
    section_fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    header_fill = openpyxl.styles.PatternFill("solid", fgColor="D9EAF7")
    thin_side = openpyxl.styles.Side(style="thin", color="D0D7DE")
    border = openpyxl.styles.Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def _section_header(cell_ref: str, title: str) -> None:
        ws[cell_ref] = title
        ws[cell_ref].font = section_font
        ws[cell_ref].fill = section_fill

    def _write_two_col_table(start_row: int, title: str, rows: list[tuple[str, object]]) -> int:
        _section_header(f"A{start_row}", title)
        ws[f"A{start_row + 1}"] = "Metric"
        ws[f"B{start_row + 1}"] = "Value"
        ws[f"A{start_row + 1}"].font = bold
        ws[f"B{start_row + 1}"].font = bold
        ws[f"A{start_row + 1}"].fill = header_fill
        ws[f"B{start_row + 1}"].fill = header_fill
        current_row = start_row + 2
        for label, value in rows:
            ws[f"A{current_row}"] = label
            ws[f"B{current_row}"] = value
            current_row += 1
        return current_row

    def _apply_formats(cell_ref: str, label: str, value: object) -> None:
        if value is None:
            return
        if isinstance(value, (int, float)):
            if any(token in label for token in ("Price", "Percentile", "Target", "Mean", "Median", "Std Dev")):
                ws[cell_ref].number_format = money_fmt
            elif any(token in label for token in ("Upside", "Probability", "Percentile Rank")):
                ws[cell_ref].number_format = pct_fmt
            elif any(token in label for token in ("Iterations", "Draws", "Samples")):
                ws[cell_ref].number_format = int_fmt

    ws["A1"] = "Monte Carlo Valuation Dashboard"
    ws["A1"].font = title_font
    ws["A2"] = f"Workbook: {path.name}"
    ws["A3"] = f"Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    summary_rows = [
        ("Current Price", results.get("current_price")),
        ("Base DCF Target", results.get("target_price")),
        ("Target Upside / Downside", results.get("target_upside")),
        ("Monte Carlo Mean", results.get("mean")),
        ("Monte Carlo Median", results.get("median")),
        ("Monte Carlo Mean Upside / Downside", results.get("mean_upside")),
        ("Std Dev", results.get("std")),
        ("Probability Price > Current", results.get("prob_above_current")),
        ("Current Price Percentile Rank", results.get("current_price_percentile")),
        ("Iterations", results.get("n_iterations")),
        ("Valid Draws", results.get("n_valid")),
        ("Sobol Samples", results.get("sobol_samples")),
        ("Runtime (s)", results.get("elapsed_seconds")),
    ]
    percentiles_rows = [
        ("5th Percentile", results.get("p5")),
        ("25th Percentile", results.get("p25")),
        ("50th Percentile", results.get("p50")),
        ("75th Percentile", results.get("p75")),
        ("95th Percentile", results.get("p95")),
    ]
    next_row = _write_two_col_table(5, "Summary", summary_rows)
    next_row = _write_two_col_table(next_row + 1, "Distribution Percentiles", percentiles_rows)

    for row in range(6, next_row):
        for col in ("A", "B"):
            ws[f"{col}{row}"].border = border
            if col == "B":
                _apply_formats(f"{col}{row}", str(ws[f"A{row}"].value or ""), ws[f"{col}{row}"].value)

    sobol_df = results.get("sobol_df")
    if sobol_df is not None:
        start_col = 4
        start_row = 5
        ws.cell(row=start_row, column=start_col, value="Sensitivity Ranking").font = section_font
        ws.cell(row=start_row, column=start_col).fill = section_fill
        for offset, header in enumerate(sobol_df.columns, start=0):
            cell = ws.cell(row=start_row + 1, column=start_col + offset, value=header)
            cell.font = bold
            cell.fill = header_fill
            cell.border = border
        for row_idx, row_data in enumerate(sobol_df.itertuples(index=False), start=start_row + 2):
            for col_idx, value in enumerate(row_data, start=start_col):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if isinstance(value, float) and sobol_df.columns[col_idx - start_col] != "Parameter":
                    cell.number_format = "0.000"

    tornado_df = results.get("tornado_df")
    if tornado_df is not None:
        start_col = 4
        start_row = 20
        tornado_cols = ["Label", "Shock Used", "Bear Delta", "Bull Delta", "Bear Price", "Bull Price"]
        ws.cell(row=start_row, column=start_col, value="Tornado Sensitivity ($/share)").font = section_font
        ws.cell(row=start_row, column=start_col).fill = section_fill
        for offset, header in enumerate(tornado_cols, start=0):
            cell = ws.cell(row=start_row + 1, column=start_col + offset, value=header)
            cell.font = bold
            cell.fill = header_fill
            cell.border = border
        for row_idx, row_data in enumerate(tornado_df[tornado_cols].itertuples(index=False), start=start_row + 2):
            for col_idx, value in enumerate(row_data, start=start_col):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if isinstance(value, float):
                    cell.number_format = money_fmt

    if chart_paths:
        image_specs = [
            ("distribution", "J2", 880, 460),
            ("percentiles", "J28", 640, 320),
            ("sobol", "J46", 820, 420),
            ("tornado", "J70", 860, 460),
        ]
        for key, anchor, width, height in image_specs:
            image_path = chart_paths.get(key)
            if image_path is None:
                continue
            image_file = Path(image_path)
            if not image_file.exists():
                continue
            image = ExcelImage(str(image_file))
            image.width = width
            image.height = height
            ws.add_image(image, anchor)

    ws.freeze_panes = "A6"
    for column, width in {
        "A": 34,
        "B": 18,
        "D": 24,
        "E": 16,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
    }.items():
        ws.column_dimensions[column].width = width

    wb.save(filepath)
    print(f"  Results written to: {filepath} [Results sheet]")


# ---------------------------------------------------------------------------
# Main — create stub, round-trip test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("=== Phase 6: Excel I/O ===\n")

    create_stub_excel("model_inputs.xlsx")
    inputs = read_inputs_from_excel("model_inputs.xlsx")
    print(f"  Read {len(inputs)} parameters:")
    for k, v in inputs.items():
        print(f"    {k:>25s} = {v}")

    # Test DCF sheet round-trip
    assert has_dcf_sheet("model_inputs.xlsx"), "DCF sheet should exist"
    dcf_data = read_dcf_from_excel("model_inputs.xlsx")
    print(f"\n  DCF sheet: {dcf_data['n_years']} years, {len(dcf_data['arrays'])} line items")
    for item, arr in dcf_data["arrays"].items():
        print(f"    {item:>15s}: {arr}")
    print(f"    WACC = {dcf_data['wacc']}, Exit Multiple = {dcf_data['exit_multiple']}")

    # Write dummy results
    dummy_results = {
        "mean": 81.0,
        "median": 64.72,
        "std": 57.40,
        "p5": 19.49,
        "p25": 39.58,
        "p50": 64.72,
        "p75": 105.59,
        "p95": 201.70,
        "n_iterations": 50000,
        "n_valid": 50000,
        "elapsed_seconds": 0.023,
    }
    write_results_to_excel("model_inputs.xlsx", dummy_results)
    print("\n  Round-trip test passed.")
